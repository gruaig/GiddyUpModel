#!/usr/bin/env python3
"""
Generate formatted Excel betting report from HorseBot CSV logs.

Usage:
    python3 generate_betting_report.py 2025-10-18
    
Creates Excel file with:
- Green highlighting for matched bets
- PNL tracking columns
- Running total
- Daily summary
"""

import sys
import csv
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("\n" + "="*80)
    print("Missing openpyxl! Install with:")
    print("  pip3 install openpyxl")
    print("="*80 + "\n")
    sys.exit(1)


def create_betting_report(date: str):
    """Create formatted Excel report from CSV logs."""
    
    log_dir = Path(__file__).parent / "strategies" / "logs" / "automated_bets"
    
    # Input files
    action_csv = log_dir / f"bot_actions_{date}.csv"
    price_csv = log_dir / f"bot_prices_{date}.csv"
    
    # Output file
    excel_file = log_dir / f"betting_report_{date}.xlsx"
    
    if not action_csv.exists():
        print(f"❌ No action log found: {action_csv}")
        return
    
    print(f"📊 Generating betting report for {date}...")
    print(f"   Reading: {action_csv}")
    
    # Create workbook
    wb = Workbook()
    
    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 1: BETTING LOG
    # ═══════════════════════════════════════════════════════════════════════════
    
    ws_bets = wb.active
    ws_bets.title = "Betting Log"
    
    # Define colors
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Light red
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Light yellow
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") # Blue
    
    bold_font = Font(bold=True, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Time", "Course", "Horse", "Strategy", "Expected", "Min", "Actual",
        "Stake", "Bet?", "Bet ID", "Reason", "Result", "Missed?", "PNL", "Running PNL"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws_bets.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Read CSV and populate
    with action_csv.open() as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        
        for row_idx, row in enumerate(rows, 2):
            # Data
            ws_bets.cell(row_idx, 1, row["race_time"])
            ws_bets.cell(row_idx, 2, row["course"])
            ws_bets.cell(row_idx, 3, row["horse"])
            ws_bets.cell(row_idx, 4, row["strategy"])
            ws_bets.cell(row_idx, 5, float(row["expected_odds"]))
            ws_bets.cell(row_idx, 6, float(row["min_odds"]))
            ws_bets.cell(row_idx, 7, float(row["actual_odds"]) if row["actual_odds"] else "")
            ws_bets.cell(row_idx, 8, float(row["stake"]))
            ws_bets.cell(row_idx, 9, row["bet_placed"])
            ws_bets.cell(row_idx, 10, row["bet_id"])
            ws_bets.cell(row_idx, 11, row["reason"])
            ws_bets.cell(row_idx, 12, "")  # Result (manual entry: WIN/LOSS for placed bets)
            ws_bets.cell(row_idx, 13, "")  # Missed Winner (manual entry: YES if would have won but didn't bet)
            
            # PNL formula: Only count if bet was actually placed
            # WIN: (Actual Odds × Stake) - Stake - (Actual Odds × Stake × 0.02) = Profit - Commission
            # LOSS: -Stake
            pnl_formula = f'=IF(I{row_idx}="NO","",IF(L{row_idx}="WIN",G{row_idx}*H{row_idx}-H{row_idx}-G{row_idx}*H{row_idx}*0.02,IF(L{row_idx}="LOSS",-H{row_idx},"")))'
            ws_bets.cell(row_idx, 14, pnl_formula)
            
            # Running PNL: Sum of all PNL above (only actual bets)
            if row_idx == 2:
                running_formula = f"=IF(ISNUMBER(N{row_idx}),N{row_idx},0)"
            else:
                running_formula = f"=IF(ISNUMBER(N{row_idx}),O{row_idx-1}+N{row_idx},O{row_idx-1})"
            ws_bets.cell(row_idx, 15, running_formula)
            
            # Apply colors
            for col in range(1, 16):
                cell = ws_bets.cell(row_idx, col)
                cell.border = thin_border
                
                # Green for matched bets
                if row["bet_placed"] in ["DRY_RUN", "EXECUTED", "YES"]:
                    cell.fill = green_fill
                # Red for missed winners (will be filled in manually)
                # Yellow for skipped (but had criteria)
                elif "too low" not in row["reason"].lower():
                    cell.fill = yellow_fill
    
    # Column widths
    ws_bets.column_dimensions['A'].width = 8   # Time
    ws_bets.column_dimensions['B'].width = 15  # Course
    ws_bets.column_dimensions['C'].width = 25  # Horse
    ws_bets.column_dimensions['D'].width = 15  # Strategy
    ws_bets.column_dimensions['E'].width = 10  # Expected
    ws_bets.column_dimensions['F'].width = 8   # Min
    ws_bets.column_dimensions['G'].width = 10  # Actual
    ws_bets.column_dimensions['H'].width = 8   # Stake
    ws_bets.column_dimensions['I'].width = 12  # Bet?
    ws_bets.column_dimensions['J'].width = 20  # Bet ID
    ws_bets.column_dimensions['K'].width = 30  # Reason
    ws_bets.column_dimensions['L'].width = 10  # Result
    ws_bets.column_dimensions['M'].width = 10  # Missed?
    ws_bets.column_dimensions['N'].width = 10  # PNL
    ws_bets.column_dimensions['O'].width = 12  # Running PNL
    
    # Freeze header row
    ws_bets.freeze_panes = 'A2'
    
    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 2: DAILY SUMMARY
    # ═══════════════════════════════════════════════════════════════════════════
    
    ws_summary = wb.create_sheet("Daily Summary")
    
    # Title
    ws_summary['A1'] = f"🏇 Betting Summary - {date}"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    ws_summary.merge_cells('A1:D1')
    
    # Stats
    bets_placed = sum(1 for r in rows if r["bet_placed"] in ["DRY_RUN", "EXECUTED", "YES"])
    bets_skipped = len(rows) - bets_placed
    total_staked = sum(float(r["stake"]) for r in rows if r["bet_placed"] in ["DRY_RUN", "EXECUTED", "YES"])
    
    summary_data = [
        ("", ""),
        ("", ""),
        ("📋 Overview", ""),
        ("Total Selections", len(rows)),
        ("Bets Placed", bets_placed),
        ("Bets Skipped", bets_skipped),
        ("", ""),
        ("💰 Stakes", ""),
        ("Total Staked", f"£{total_staked:.2f}"),
        ("Average Stake", f"£{total_staked/bets_placed if bets_placed > 0 else 0:.2f}"),
        ("", ""),
        ("📊 Results (Fill in manually)", ""),
        ("Wins", "=COUNTIF('Betting Log'!L:L,\"WIN\")"),
        ("Losses", "=COUNTIF('Betting Log'!L:L,\"LOSS\")"),
        ("Pending", f"={bets_placed}-B14-B15"),
        ("Missed Winners", "=COUNTIF('Betting Log'!M:M,\"YES\")"),
        ("", ""),
        ("💵 Profit & Loss", ""),
        ("Total PNL", "=IFERROR('Betting Log'!O" + str(len(rows)+1) + ",0)"),
        ("ROI", "=IFERROR(B20/B10*100,0)&\"%\""),
        ("Win Rate", "=IFERROR(B14/(B14+B15)*100,0)&\"%\""),
        ("Missed Win Rate", "=IFERROR(B17/(B17+B15)*100,0)&\"%\""),
        ("", ""),
        ("🎯 Strategy Breakdown", ""),
        ("Strategy A Bets", "=COUNTIFS('Betting Log'!D:D,\"A-Hybrid_V3\",'Betting Log'!I:I,\"DRY_RUN\")+COUNTIFS('Betting Log'!D:D,\"A-Hybrid_V3\",'Betting Log'!I:I,\"EXECUTED\")"),
        ("Strategy B Bets", "=COUNTIFS('Betting Log'!D:D,\"B-Path_B\",'Betting Log'!I:I,\"DRY_RUN\")+COUNTIFS('Betting Log'!D:D,\"B-Path_B\",'Betting Log'!I:I,\"EXECUTED\")"),
    ]
    
    # Start from row 2 to avoid merged cell
    for idx, (label, value) in enumerate(summary_data, 2):
        ws_summary.cell(idx, 1, label).font = Font(bold=True)
        if value:
            ws_summary.cell(idx, 2, value)
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 3: PRICE MOVEMENTS (if available)
    # ═══════════════════════════════════════════════════════════════════════════
    
    if price_csv.exists():
        ws_prices = wb.create_sheet("Price Movements")
        
        # Headers
        price_headers = ["Timestamp", "Race Time", "Course", "Horse", "T-Mins", "Odds", "Status"]
        for col, header in enumerate(price_headers, 1):
            cell = ws_prices.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        
        # Read price data
        with price_csv.open() as f:
            reader = csv.DictReader(f)
            for row_idx, row in enumerate(reader, 2):
                ws_prices.cell(row_idx, 1, row["timestamp"])
                ws_prices.cell(row_idx, 2, row["race_time"])
                ws_prices.cell(row_idx, 3, row["course"])
                ws_prices.cell(row_idx, 4, row["horse"])
                ws_prices.cell(row_idx, 5, float(row["minutes_to_off"]) if row["minutes_to_off"] else "")
                ws_prices.cell(row_idx, 6, float(row["odds"]) if row["odds"] else "")
                ws_prices.cell(row_idx, 7, row["status"])
                
                # Color code by status
                for col in range(1, 8):
                    cell = ws_prices.cell(row_idx, col)
                    cell.border = thin_border
                    if row["status"] == "BET_WINDOW":
                        cell.fill = yellow_fill
        
        ws_prices.column_dimensions['A'].width = 20
        ws_prices.column_dimensions['B'].width = 10
        ws_prices.column_dimensions['C'].width = 15
        ws_prices.column_dimensions['D'].width = 25
        ws_prices.column_dimensions['E'].width = 10
        ws_prices.column_dimensions['F'].width = 10
        ws_prices.column_dimensions['G'].width = 15
        
        ws_prices.freeze_panes = 'A2'
    
    # ═══════════════════════════════════════════════════════════════════════════
    # SAVE
    # ═══════════════════════════════════════════════════════════════════════════
    
    wb.save(excel_file)
    
    print(f"✅ Report generated: {excel_file}")
    print("")
    print("📝 Next steps:")
    print(f"   1. Open in Excel/LibreOffice: {excel_file}")
    print(f"   2. Fill in 'Result' column (L): Enter 'WIN' or 'LOSS'")
    print(f"   3. PNL will auto-calculate")
    print(f"   4. Check 'Daily Summary' sheet for totals")
    print("")
    print("💡 Green rows = Bets placed/executed")
    print("💡 Yellow rows = Market found but skipped")
    print("")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("\nUsage: python3 generate_betting_report.py <date>")
        print("Example: python3 generate_betting_report.py 2025-10-18\n")
        sys.exit(1)
    
    date = sys.argv[1]
    create_betting_report(date)

